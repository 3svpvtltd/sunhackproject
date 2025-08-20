import re
import google.generativeai as genai
from flask import Flask, request, jsonify, render_template
import os
from dotenv import load_dotenv
from gtts import gTTS
from flask import send_from_directory
import openpyxl
import smtplib
from email.mime.text import MIMEText

# Initialize Flask app
app = Flask(__name__)

# Load environment variables
load_dotenv()

# Configure Gemini API with the key from environment variable
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

# Define the generation configuration
generation_config = {
    "temperature": 1,
    "top_p": 0.95,
    "top_k": 40,
    "max_output_tokens": 8192,
    "response_mime_type": "text/plain",
}

# Create the Gemini model
model = genai.GenerativeModel(
    model_name="gemini-2.0-flash",
    generation_config=generation_config,
)

# Start a chat session
chat_session = model.start_chat(history=[])

# Store the last AI response globally
last_response = {"text": ""}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/text')
def text():
    return render_template("text.html")

@app.route('/game1')
def game1():
    return render_template("game1.html")

@app.route('/game2')
def game2():
    return render_template("game2.html")

@app.route('/game3')
def game3():
    return render_template("game3.html")

@app.route('/game4')
def game4():
    return render_template("game4.html")

@app.route('/game5')
def game5():
    return render_template("game5.html")

@app.route('/pz')
def pz():
    return render_template("pz.html")

@app.route('/loader')
def loader():
    return render_template("loader.html")
# @app.route('/getdata', methods=['POST'])
# def getdata():
#     name = request.form.get('name')
#     email = request.form.get('email')
#     message = request.form.get('message')

#     if not email:  # Check if email is None or empty
#         return "Error: Email address is required!", 400

#     # Define the Excel file
#     excel_file = "form_data.xlsx"

#     try:
#         workbook = openpyxl.load_workbook(excel_file)  
#         sheet = workbook.active
#     except FileNotFoundError:
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         sheet.append(["name", "email", "message"]) 

#     sheet.append([name, email, message])
#     workbook.save(excel_file)  

#     try:
#         server = smtplib.SMTP('smtp.gmail.com', 587)  # Corrected port
#         server.starttls()
#         server.login('saichechare63@gmail.com', 'tsbs dspp rjln quud')  # Use App Password instead of actual password
#         server.sendmail(
#             'saichechare63@gmail.com', 
#             email, 
#             f'Subject: Thank You!\n\nThank you for sending a message. You will get a reply as soon as possible...\nYour Subject: {message}'
#         )
#         server.quit()
#     except Exception as e:
#         return f"Error sending email: {e}", 500

#     return render_template('index.html')
@app.route('/gemini-response', methods=['POST'])
def get_gemini_response():
    global last_response  # Access the global variable

    try:
        user_message = request.json['message']
        
        if not user_message:
            return jsonify({"error": "Message is required"}), 400

        response = chat_session.send_message(user_message)
        ai_response = response.text.strip()
        formatted_response = format_response(ai_response)

        # Store the response in a global variable
        last_response["text"] = formatted_response  

        return jsonify({"response": formatted_response})

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal Server Error"}), 500

def format_response(response_text):
    response_text = re.sub(r'\*\*(.*?)\*\*', r'\1', response_text)  
    response_text = re.sub(r'__(.*?)__', r'\1', response_text)  
    response_text = re.sub(r'\s+', ' ', response_text)  

    tts = gTTS(text=response_text, lang='en', lang_check=True)
    tts.save(r'E:\\new hackthone shegoan\\static\\audio.mp3')

    return response_text

@app.route("/start")
def start():
    if not last_response["text"]:
        return "No topic available for MCQ generation!", 400

    new_message = f"Generate 5 multiple-choice questions with 4 options each. " \
                  f"Mark the correct answer by appending '(Correct)' to it:\n{last_response['text']}"

    try:
        # Send the request to Gemini
        response = chat_session.send_message(new_message)
        raw_mcqs = response.text.strip()

        if not raw_mcqs:
            return "Failed to generate MCQs. Try again later.", 500

        # Process MCQs to ensure correct answers are marked
        formatted_mcqs = format_mcq_output(raw_mcqs)

        # Store the new response
        last_response["text"] = formatted_mcqs

        return render_template("start.html", mcq_response=formatted_mcqs)

    except Exception as e:
        print(f"Error generating MCQs: {e}")
        return "Error generating MCQs. Please try again.", 500

def format_mcq_output(mcq_text):
    """
    Formats the MCQs properly, ensuring correct answers are marked.
    """
    questions = mcq_text.split("\n\n")
    formatted_questions = []

    for question in questions:
        lines = question.split("\n")
        if len(lines) < 2:
            continue  # Skip if the format is incorrect

        # Ensure the correct answer is marked
        question_text = lines[0]
        options = lines[1:]

        found_correct = False
        for i in range(len(options)):
            if "*" in options[i] or "**" in options[i]:  # Check if Gemini marks correct answers
                options[i] = options[i].replace("*", "").strip() + " (Correct)"
                found_correct = True

        if not found_correct and options:  
            options[0] += " (Correct)"  # Mark the first option correct if none is marked

        formatted_questions.append("\n".join([question_text] + options))

    return "\n\n".join(formatted_questions)


@app.route('/getdata', methods=['POST'])
def getdata():
    parent_name = request.form.get("parent_name")
    email = request.form.get("email")
    child_age = request.form.get("child_age")
    interest = request.form.get("interest")
    message = request.form.get("message")

    if not email:
        return "Error: Email address is required!", 400

    excel_file = "form_data.xlsx"
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Parent Name", "Email", "Child Age", "Interest", "Message"])

    sheet.append([parent_name, email, child_age, interest, message])
    workbook.save(excel_file)

    # Send email
    try:
        msg_content = (
            f"Hello {parent_name},\n\n"
            f"Thank you for reaching out!\n"
            f"Child Age: {child_age}\n"
            f"Interest: {interest}\n"
            f"Message: {message}\n\n"
            f"We’ll get back to you soon.\n\n"
            f"Best Regards,\nTeam"
        )
        msg = MIMEText(msg_content, "plain", "utf-8")
        msg["Subject"] = "Thank You!"
        msg["From"] = "saichechare63@gmail.com"
        msg["To"] = email

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login("saichechare63@gmail.com", "tsbs dspp rjln quud")
        server.send_message(msg)  # ✅ Correct usage
        server.quit()
    except Exception as e:
        return f"Error sending email: {e}", 500

    # ✅ Render home page after email success
    return render_template("index.html")


if __name__ == '__main__':
    app.run(debug=True)